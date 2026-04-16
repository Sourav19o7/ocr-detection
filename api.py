"""
FastAPI OCR API - Hallmark QC Validation System.

Three-Stage Workflow:
1. Stage 1: Upload CSV/Excel with tag IDs and expected HUIDs
2. Stage 2: Upload images with tag ID for processing
3. Stage 3: Retrieve results by tag ID

Deploy on AWS or run locally with: uvicorn api:app --host 0.0.0.0 --port 8000
"""

from fastapi import FastAPI, File, UploadFile, HTTPException, Form, Query, Request, Depends
from fastapi.middleware.cors import CORSMiddleware
from fastapi.staticfiles import StaticFiles
from fastapi.responses import JSONResponse, RedirectResponse, FileResponse
from pydantic import BaseModel
from typing import Optional, List, Dict
from PIL import Image
from datetime import datetime
from starlette.middleware.sessions import SessionMiddleware
import io
import sys
import os
import time
import pandas as pd
import boto3
from botocore.exceptions import ClientError
from dotenv import load_dotenv
import secrets

# Load environment variables
load_dotenv()

# Import OCR engines
sys.path.insert(0, "src")
sys.path.insert(0, "config")

from ocr_model import OCREngine, OCRResult
from ocr_model_v2 import OCREngineV2, OCRResultV2, HallmarkInfo, HallmarkType, CheckInfo
from database import (
    DatabaseManager, get_database, Batch, BatchItem, OCRResult as DBOCRResult,
    ProcessingStatus, QCDecision
)
from storage_service import StorageService, get_storage


app = FastAPI(
    title="Hallmark QC API",
    description="""
    AI-powered Hallmark Quality Control System with BIS compliance validation.

    ## Three-Stage Workflow

    ### Stage 1: Batch Upload
    Upload CSV/Excel files containing tag IDs and expected HUIDs.

    ### Stage 2: Image Processing
    Upload hallmark images with their corresponding tag IDs for OCR processing.

    ### Stage 3: Results Retrieval
    Retrieve OCR results, expected vs actual HUID comparison, and processed images.

    ## BIS Standards
    - Gold: IS 1417 (916, 750, 585, etc.)
    - Silver: IS 2112 (925, 999, etc.)
    - HUID: Mandatory 6-character alphanumeric code (since April 2023)
    """,
    version="4.0.0"
)

# Session middleware for authentication
app.add_middleware(
    SessionMiddleware,
    secret_key=os.getenv('SESSION_SECRET', secrets.token_hex(32))
)

# CORS middleware
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Parse login users from environment variable (format: user1:pass1,user2:pass2)
login_users = {}
if os.getenv('LOGIN_USERS'):
    for user_pass in os.getenv('LOGIN_USERS').split(','):
        parts = user_pass.strip().split(':')
        if len(parts) == 2:
            login_users[parts[0].strip()] = parts[1].strip()

# Mount uploads directory for serving local images
uploads_dir = "./uploads"
os.makedirs(uploads_dir, exist_ok=True)
app.mount("/uploads", StaticFiles(directory=uploads_dir), name="uploads")

# Initialize OCR engines at startup
ocr_engine = None
ocr_engine_v2 = None
db: DatabaseManager = None
storage: StorageService = None

# S3 Client for presigned URLs
s3_client = None

def get_s3_client():
    """Get or create S3 client."""
    global s3_client
    if s3_client is None:
        s3_client = boto3.client(
            's3',
            region_name=os.getenv('AWS_REGION', 'ap-south-1'),
            aws_access_key_id=os.getenv('AWS_ACCESS_KEY_ID'),
            aws_secret_access_key=os.getenv('AWS_SECRET_ACCESS_KEY')
        )
    return s3_client


# Response Models
class BatchUploadResponse(BaseModel):
    status: str
    batch_id: int
    batch_name: str
    total_items: int
    message: str


class ImageUploadResponse(BaseModel):
    status: str
    tag_id: str
    expected_huid: str
    actual_huid: Optional[str]
    huid_match: bool
    confidence: float
    decision: str
    message: str


class ResultResponse(BaseModel):
    status: str
    tag_id: str
    expected_huid: str
    actual_huid: Optional[str]
    huid_match: Optional[bool]
    purity_code: Optional[str]
    karat: Optional[str]
    purity_percentage: Optional[float]
    confidence: Optional[float]
    decision: Optional[str]
    rejection_reasons: List[str]
    raw_ocr_text: Optional[str]
    image_url: Optional[str]
    processed_image_url: Optional[str]
    processing_status: str


class BatchResultsResponse(BaseModel):
    status: str
    batch_id: int
    batch_name: str
    total_items: int
    processed_items: int
    statistics: Dict
    results: List[Dict]


class HallmarkData(BaseModel):
    purity_code: Optional[str] = None
    karat: Optional[str] = None
    purity_percentage: Optional[float] = None
    huid: Optional[str] = None
    bis_certified: bool = False


@app.on_event("startup")
async def startup_event():
    """Initialize services on startup."""
    global ocr_engine, ocr_engine_v2, db, storage
    ocr_engine = OCREngine()
    ocr_engine_v2 = OCREngineV2(enable_preprocessing=True)
    db = get_database()
    storage = get_storage()
    print(f"Storage type: {storage.storage_type}")


# =============================================================================
# Homepage & Health Endpoints
# =============================================================================

@app.get("/", response_class=RedirectResponse)
async def homepage():
    """Redirect to API documentation."""
    return RedirectResponse(url="/docs")


@app.get("/api/health")
async def api_health():
    """API health check endpoint."""
    return {
        "status": "ok",
        "message": "Hallmark QC API is running",
        "version": "4.0.0",
        "storage": storage.storage_type if storage else "not_initialized"
    }


@app.get("/health")
async def health():
    """Health check for deployment."""
    return {"status": "healthy"}


# =============================================================================
# STAGE 1: Batch Upload (CSV/Excel with tag IDs and expected HUIDs)
# =============================================================================

@app.post("/stage1/upload-batch", response_model=BatchUploadResponse)
async def upload_batch(
    file: UploadFile = File(...),
    batch_name: Optional[str] = Form(None)
):
    """
    Stage 1: Upload CSV/Excel file with tag IDs and expected HUIDs.

    The file should have at least two columns:
    - tag_id: Unique identifier for each item
    - expected_huid: The expected HUID to validate against

    Supported formats: CSV, XLSX, XLS
    """
    # Validate file type
    allowed_types = [
        "text/csv",
        "application/vnd.ms-excel",
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    ]
    filename_lower = file.filename.lower()

    if file.content_type not in allowed_types and not (
        filename_lower.endswith(".csv") or
        filename_lower.endswith(".xlsx") or
        filename_lower.endswith(".xls")
    ):
        raise HTTPException(
            status_code=400,
            detail="Invalid file type. Allowed: CSV, XLSX, XLS"
        )

    try:
        contents = await file.read()

        # Parse file based on type
        if filename_lower.endswith(".csv"):
            df = pd.read_csv(io.BytesIO(contents))
        else:
            # For Excel files, try multiple methods to handle compatibility issues
            try:
                # Try with openpyxl first (for .xlsx files)
                excel_file = pd.ExcelFile(io.BytesIO(contents), engine='openpyxl')
                sheet_names = excel_file.sheet_names

                # Try to find a sheet with 'HUID' in the name
                huid_sheet = None
                for sheet in sheet_names:
                    if 'HUID' in sheet.upper() or 'PRINT' in sheet.upper():
                        huid_sheet = sheet
                        break

                # Read the HUID sheet if found, otherwise read the first sheet
                if huid_sheet:
                    df = pd.read_excel(io.BytesIO(contents), sheet_name=huid_sheet, engine='openpyxl')
                else:
                    df = pd.read_excel(io.BytesIO(contents), sheet_name=0, engine='openpyxl')

            except Exception as openpyxl_error:
                # If openpyxl fails due to styling issues, read with data_only mode
                try:
                    from openpyxl import load_workbook
                    wb = load_workbook(io.BytesIO(contents), data_only=True, read_only=True)

                    # Find HUID sheet
                    huid_sheet_name = None
                    for sheet_name in wb.sheetnames:
                        if 'HUID' in sheet_name.upper() or 'PRINT' in sheet_name.upper():
                            huid_sheet_name = sheet_name
                            break

                    # Use the HUID sheet or first sheet
                    ws = wb[huid_sheet_name] if huid_sheet_name else wb[wb.sheetnames[0]]

                    # Convert to dataframe
                    data = []
                    for row in ws.iter_rows(values_only=True):
                        if row and any(cell is not None for cell in row):  # Skip empty rows
                            data.append(row)

                    if len(data) < 2:  # Need at least header + 1 data row
                        raise HTTPException(status_code=400, detail="Excel file has insufficient data")

                    df = pd.DataFrame(data[1:], columns=data[0])

                except Exception as fallback_error:
                    raise HTTPException(
                        status_code=400,
                        detail=f"Failed to read Excel file. Try saving as a new .xlsx file or export to CSV. Original error: {str(openpyxl_error)}"
                    )

        # Normalize column names
        df.columns = df.columns.str.lower().str.strip().str.replace(" ", "_")

        # Validate required columns
        required_columns = ["tag_id", "expected_huid"]
        # Also check for common variations
        column_mappings = {
            "tag_id": ["tag_id", "tagid", "tag", "id", "item_id"],
            "expected_huid": ["expected_huid", "huid", "expected_id", "expectedhuid"]
        }

        for req_col, variations in column_mappings.items():
            found = False
            for var in variations:
                if var in df.columns:
                    if var != req_col:
                        df = df.rename(columns={var: req_col})
                    found = True
                    break
            if not found:
                raise HTTPException(
                    status_code=400,
                    detail=f"Missing required column: {req_col}. Found columns: {list(df.columns)}"
                )

        # Remove duplicates and empty rows
        df = df.dropna(subset=["tag_id", "expected_huid"])
        df = df.drop_duplicates(subset=["tag_id"])

        if len(df) == 0:
            raise HTTPException(
                status_code=400,
                detail="No valid rows found in the file"
            )

        # Create batch
        batch_name = batch_name or f"Batch_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
        batch = Batch(
            batch_name=batch_name,
            total_items=len(df),
            status="pending"
        )
        batch_id = db.create_batch(batch)

        # Create batch items
        for _, row in df.iterrows():
            item = BatchItem(
                batch_id=batch_id,
                tag_id=str(row["tag_id"]).strip(),
                expected_huid=str(row["expected_huid"]).strip().upper(),
                status=ProcessingStatus.PENDING
            )
            db.create_batch_item(item)

        return BatchUploadResponse(
            status="success",
            batch_id=batch_id,
            batch_name=batch_name,
            total_items=len(df),
            message=f"Successfully uploaded {len(df)} items"
        )

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to process file: {str(e)}")


@app.get("/stage1/batches")
async def list_batches():
    """List all uploaded batches."""
    batches = db.get_all_batches()
    return {
        "status": "success",
        "batches": [b.to_dict() for b in batches]
    }


@app.get("/stage1/batch/{batch_id}")
async def get_batch_details(batch_id: int):
    """Get details of a specific batch."""
    batch = db.get_batch(batch_id)
    if not batch:
        raise HTTPException(status_code=404, detail="Batch not found")

    items = db.get_batch_items(batch_id)
    stats = db.get_batch_statistics(batch_id)

    return {
        "status": "success",
        "batch": batch.to_dict(),
        "statistics": stats,
        "items": [item.to_dict() for item in items]
    }


# =============================================================================
# STAGE 2: Image Upload and Processing
# =============================================================================

@app.post("/stage2/upload-image", response_model=ImageUploadResponse)
async def upload_and_process_image(
    file: UploadFile = File(...),
    tag_id: str = Form(...)
):
    """
    Stage 2: Upload an image for a specific tag ID and process it.

    The image will be:
    1. Stored in S3 (or local storage)
    2. Processed through OCR
    3. Compared against the expected HUID from Stage 1
    """
    # Validate file type
    allowed_types = ["image/png", "image/jpeg", "image/jpg", "image/bmp", "image/webp"]
    if file.content_type not in allowed_types:
        raise HTTPException(
            status_code=400,
            detail="Invalid file type. Allowed: PNG, JPG, JPEG, BMP, WEBP"
        )

    # Check if tag_id exists in database
    batch_item = db.get_batch_item_by_tag(tag_id)
    if not batch_item:
        raise HTTPException(
            status_code=404,
            detail=f"Tag ID '{tag_id}' not found. Please upload batch data first (Stage 1)."
        )

    try:
        start_time = time.time()

        # Read and validate image
        contents = await file.read()
        image = Image.open(io.BytesIO(contents))

        if image.mode != "RGB":
            image = image.convert("RGB")

        # Upload original image to storage
        image_path, image_url = storage.upload_image(
            contents,
            tag_id,
            file.filename,
            prefix="originals"
        )

        # Update batch item with image path
        db.update_batch_item_image(tag_id, image_path, image_url)

        # Process with OCR
        hallmark_info = ocr_engine_v2.extract_with_hallmark_info(image)

        # Extract actual HUID
        actual_huid = hallmark_info.huid
        expected_huid = batch_item.expected_huid

        # Compare HUIDs
        huid_match = False
        if actual_huid and expected_huid:
            huid_match = actual_huid.upper().strip() == expected_huid.upper().strip()

        # Determine decision
        decision = QCDecision.PENDING
        rejection_reasons = []

        purity_valid = hallmark_info.purity_code is not None
        huid_valid = actual_huid is not None

        if not purity_valid:
            rejection_reasons.append("missing_purity_mark")
        if not huid_valid:
            rejection_reasons.append("missing_huid")
        if actual_huid and not huid_match:
            rejection_reasons.append("huid_mismatch")

        # Decision logic
        if purity_valid and huid_valid and huid_match:
            if hallmark_info.overall_confidence >= 0.85:
                decision = QCDecision.APPROVED
            elif hallmark_info.overall_confidence >= 0.50:
                decision = QCDecision.MANUAL_REVIEW
            else:
                decision = QCDecision.REJECTED
                rejection_reasons.append("low_confidence")
        elif purity_valid and huid_valid and not huid_match:
            decision = QCDecision.REJECTED
        else:
            if hallmark_info.overall_confidence < 0.50:
                decision = QCDecision.REJECTED
            else:
                decision = QCDecision.MANUAL_REVIEW

        processing_time = int((time.time() - start_time) * 1000)

        # Save processed image with annotations (optional)
        processed_image_path = None
        processed_image_url = None

        # Create OCR result record
        ocr_result = DBOCRResult(
            batch_item_id=batch_item.id,
            tag_id=tag_id,
            expected_huid=expected_huid,
            actual_huid=actual_huid,
            huid_match=huid_match,
            purity_code=hallmark_info.purity_code,
            karat=hallmark_info.karat,
            purity_percentage=hallmark_info.purity_percentage,
            confidence=hallmark_info.overall_confidence,
            decision=decision,
            rejection_reasons=rejection_reasons,
            raw_ocr_text=" ".join([r.text for r in hallmark_info.all_results]),
            processed_image_path=processed_image_path,
            processed_image_url=processed_image_url,
            processing_time_ms=processing_time,
        )
        db.create_ocr_result(ocr_result)

        # Update batch item status
        db.update_batch_item_status(tag_id, ProcessingStatus.COMPLETED)

        # Update batch progress
        batch = db.get_batch(batch_item.batch_id)
        if batch:
            db.update_batch_progress(batch.id, batch.processed_items + 1)

        return ImageUploadResponse(
            status="success",
            tag_id=tag_id,
            expected_huid=expected_huid,
            actual_huid=actual_huid,
            huid_match=huid_match,
            confidence=round(hallmark_info.overall_confidence, 3),
            decision=decision.value,
            message="Image processed successfully"
        )

    except HTTPException:
        raise
    except Exception as e:
        # Update status to failed
        db.update_batch_item_status(tag_id, ProcessingStatus.FAILED)
        raise HTTPException(status_code=500, detail=f"Image processing failed: {str(e)}")


@app.post("/stage2/upload-image-bulk")
async def upload_images_bulk(
    files: List[UploadFile] = File(...),
):
    """
    Stage 2 (Bulk): Upload multiple images at once.

    Image filenames should contain the tag_id (e.g., TAG001.jpg, TAG002.png).
    """
    results = []
    errors = []

    for file in files:
        # Extract tag_id from filename
        filename = os.path.splitext(file.filename)[0]
        tag_id = filename.strip()

        try:
            # Check if tag exists
            batch_item = db.get_batch_item_by_tag(tag_id)
            if not batch_item:
                errors.append({
                    "filename": file.filename,
                    "tag_id": tag_id,
                    "error": "Tag ID not found in batch data"
                })
                continue

            # Process the image
            contents = await file.read()
            image = Image.open(io.BytesIO(contents))

            if image.mode != "RGB":
                image = image.convert("RGB")

            # Upload image
            image_path, image_url = storage.upload_image(
                contents,
                tag_id,
                file.filename,
                prefix="originals"
            )

            db.update_batch_item_image(tag_id, image_path, image_url)

            # Process OCR
            hallmark_info = ocr_engine_v2.extract_with_hallmark_info(image)
            actual_huid = hallmark_info.huid
            huid_match = (
                actual_huid and
                actual_huid.upper().strip() == batch_item.expected_huid.upper().strip()
            )

            # Simple decision
            if hallmark_info.overall_confidence >= 0.85 and huid_match:
                decision = QCDecision.APPROVED
            elif hallmark_info.overall_confidence >= 0.50:
                decision = QCDecision.MANUAL_REVIEW
            else:
                decision = QCDecision.REJECTED

            # Save result
            ocr_result = DBOCRResult(
                batch_item_id=batch_item.id,
                tag_id=tag_id,
                expected_huid=batch_item.expected_huid,
                actual_huid=actual_huid,
                huid_match=huid_match,
                purity_code=hallmark_info.purity_code,
                confidence=hallmark_info.overall_confidence,
                decision=decision,
                raw_ocr_text=" ".join([r.text for r in hallmark_info.all_results]),
            )
            db.create_ocr_result(ocr_result)
            db.update_batch_item_status(tag_id, ProcessingStatus.COMPLETED)

            results.append({
                "tag_id": tag_id,
                "status": "success",
                "huid_match": huid_match,
                "decision": decision.value
            })

        except Exception as e:
            errors.append({
                "filename": file.filename,
                "tag_id": tag_id,
                "error": str(e)
            })

    return {
        "status": "completed",
        "processed": len(results),
        "errors": len(errors),
        "results": results,
        "error_details": errors
    }


# =============================================================================
# STAGE 3: Results Retrieval
# =============================================================================

@app.get("/stage3/result/{tag_id}", response_model=ResultResponse)
async def get_result_by_tag(tag_id: str):
    """
    Stage 3: Get complete result for a specific tag ID.

    Returns:
    - Expected HUID (from batch upload)
    - Actual HUID (from OCR)
    - Match status
    - OCR confidence and decision
    - Original and processed image URLs
    """
    result = db.get_full_result_by_tag(tag_id)

    if not result:
        raise HTTPException(
            status_code=404,
            detail=f"Tag ID '{tag_id}' not found"
        )

    return ResultResponse(
        status="success",
        tag_id=result["tag_id"],
        expected_huid=result["expected_huid"],
        actual_huid=result.get("actual_huid"),
        huid_match=result.get("huid_match"),
        purity_code=result.get("purity_code"),
        karat=result.get("karat"),
        purity_percentage=result.get("purity_percentage"),
        confidence=result.get("confidence"),
        decision=result.get("decision"),
        rejection_reasons=result.get("rejection_reasons", []),
        raw_ocr_text=result.get("raw_ocr_text"),
        image_url=result.get("image_url"),
        processed_image_url=result.get("processed_image_url"),
        processing_status=result.get("status", "pending")
    )


@app.get("/stage3/batch/{batch_id}/results", response_model=BatchResultsResponse)
async def get_batch_results(batch_id: int):
    """
    Stage 3: Get all results for a batch.

    Returns complete statistics and individual results for all items in the batch.
    """
    batch = db.get_batch(batch_id)
    if not batch:
        raise HTTPException(status_code=404, detail="Batch not found")

    results = db.get_results_by_batch(batch_id)
    stats = db.get_batch_statistics(batch_id)

    return BatchResultsResponse(
        status="success",
        batch_id=batch_id,
        batch_name=batch.batch_name,
        total_items=batch.total_items,
        processed_items=stats.get("status_counts", {}).get("completed", 0),
        statistics=stats,
        results=results
    )


@app.get("/stage3/search")
async def search_results(
    tag_id: Optional[str] = Query(None, description="Search by tag ID"),
    huid: Optional[str] = Query(None, description="Search by HUID"),
    decision: Optional[str] = Query(None, description="Filter by decision"),
    batch_id: Optional[int] = Query(None, description="Filter by batch"),
):
    """Search and filter results."""
    # This would be enhanced with proper search functionality
    if tag_id:
        result = db.get_full_result_by_tag(tag_id)
        if result:
            return {"status": "success", "results": [result]}
        return {"status": "success", "results": []}

    if batch_id:
        results = db.get_results_by_batch(batch_id)
        if decision:
            results = [r for r in results if r.get("decision") == decision]
        return {"status": "success", "results": results}

    return {"status": "success", "results": [], "message": "Please provide search criteria"}


# =============================================================================
# S3 Presigned URL Endpoints (for direct browser uploads)
# =============================================================================

class PresignedUrlRequest(BaseModel):
    fileName: str
    contentType: Optional[str] = "application/octet-stream"


class PresignedUrlResponse(BaseModel):
    uploadUrl: str
    key: str
    fileName: str


@app.post("/api/get-upload-url", response_model=PresignedUrlResponse)
async def get_upload_url(request: PresignedUrlRequest):
    """
    Get a presigned URL for direct S3 upload (authenticated uploads).

    This endpoint generates a presigned URL that allows direct upload
    to S3 from the browser without routing through the server.
    """
    try:
        client = get_s3_client()
        bucket_name = os.getenv('S3_BUCKET_NAME')
        folder_path = os.getenv('S3_FOLDER_PATH', 'uploads/')

        if not bucket_name:
            raise HTTPException(status_code=500, detail="S3 bucket not configured")

        timestamp = int(time.time() * 1000)
        unique_filename = f"{timestamp}-{request.fileName}"
        s3_key = f"{folder_path}{unique_filename}"

        presigned_url = client.generate_presigned_url(
            'put_object',
            Params={
                'Bucket': bucket_name,
                'Key': s3_key,
                'ContentType': request.contentType
            },
            ExpiresIn=3600  # URL valid for 1 hour
        )

        return PresignedUrlResponse(
            uploadUrl=presigned_url,
            key=s3_key,
            fileName=unique_filename
        )
    except ClientError as e:
        raise HTTPException(status_code=500, detail=f"Failed to generate presigned URL: {str(e)}")


@app.post("/api/get-ocr-upload-url", response_model=PresignedUrlResponse)
async def get_ocr_upload_url(request: PresignedUrlRequest):
    """
    Get a presigned URL for OCR image uploads (no auth required).

    This endpoint is specifically for uploading hallmark images for OCR processing.
    Images are stored in a dedicated OCR folder in S3.
    """
    try:
        client = get_s3_client()
        bucket_name = os.getenv('S3_BUCKET_NAME')
        folder_path = os.getenv('S3_OCR_FOLDER_PATH', 'ocr-images/')

        if not bucket_name:
            raise HTTPException(status_code=500, detail="S3 bucket not configured")

        timestamp = int(time.time() * 1000)
        unique_filename = f"{timestamp}-{request.fileName}"
        s3_key = f"{folder_path}{unique_filename}"

        # Default to image/jpeg for OCR uploads
        content_type = request.contentType if request.contentType != "application/octet-stream" else "image/jpeg"

        presigned_url = client.generate_presigned_url(
            'put_object',
            Params={
                'Bucket': bucket_name,
                'Key': s3_key,
                'ContentType': content_type
            },
            ExpiresIn=3600  # URL valid for 1 hour
        )

        return PresignedUrlResponse(
            uploadUrl=presigned_url,
            key=s3_key,
            fileName=unique_filename
        )
    except ClientError as e:
        raise HTTPException(status_code=500, detail=f"Failed to generate presigned URL: {str(e)}")


# =============================================================================
# Legacy Endpoints (kept for backward compatibility)
# =============================================================================

@app.post("/extract")
async def extract_text(file: UploadFile = File(...)):
    """Legacy: Extract text from an uploaded image."""
    allowed_types = ["image/png", "image/jpeg", "image/jpg", "image/bmp", "image/webp"]
    if file.content_type not in allowed_types:
        raise HTTPException(status_code=400, detail="Invalid file type")

    try:
        contents = await file.read()
        image = Image.open(io.BytesIO(contents))

        if image.mode != "RGB":
            image = image.convert("RGB")

        ocr_results = ocr_engine.extract_text_with_confidence(image)

        results = [
            {"text": r.text, "confidence": r.confidence, "approved": r.confidence >= 0.75}
            for r in ocr_results
        ]

        return {
            "success": True,
            "results": results,
            "full_text": "\n".join([r.text for r in ocr_results]),
            "average_confidence": sum(r.confidence for r in ocr_results) / len(ocr_results) if ocr_results else 0.0
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"OCR processing failed: {str(e)}")


@app.post("/extract/v2")
async def extract_text_v2(file: UploadFile = File(...)):
    """Legacy: Extract text with hallmark detection."""
    allowed_types = ["image/png", "image/jpeg", "image/jpg", "image/bmp", "image/webp"]
    if file.content_type not in allowed_types:
        raise HTTPException(status_code=400, detail="Invalid file type")

    try:
        contents = await file.read()
        image = Image.open(io.BytesIO(contents))

        if image.mode != "RGB":
            image = image.convert("RGB")

        hallmark_info = ocr_engine_v2.extract_with_hallmark_info(image)

        # Build individual detection results
        detections = []
        for r in hallmark_info.all_results:
            detections.append({
                "text": r.text,
                "confidence": round(r.confidence, 4),
                "type": r.hallmark_type.value,
                "validated": r.validated,
            })

        return {
            "success": True,
            "full_text": " ".join([r.text for r in hallmark_info.all_results]),
            "average_confidence": hallmark_info.overall_confidence,
            "hallmark": {
                "purity_code": hallmark_info.purity_code,
                "karat": hallmark_info.karat,
                "purity_percentage": hallmark_info.purity_percentage,
                "huid": hallmark_info.huid,
                "bis_certified": hallmark_info.bis_certified
            },
            "detections": detections
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"OCR processing failed: {str(e)}")


@app.get("/qc/rules")
async def get_qc_rules():
    """Get current QC validation rules and BIS compliance standards."""
    return {
        "bis_standards": {
            "gold_grades": {
                "375": {"karat": "9K", "purity": 37.5},
                "585": {"karat": "14K", "purity": 58.5},
                "750": {"karat": "18K", "purity": 75.0},
                "875": {"karat": "21K", "purity": 87.5},
                "916": {"karat": "22K", "purity": 91.6},
                "958": {"karat": "23K", "purity": 95.8},
                "999": {"karat": "24K", "purity": 99.9},
            },
            "silver_grades": {
                "800": {"purity": 80.0, "grade": "Silver 800"},
                "925": {"purity": 92.5, "grade": "Sterling Silver"},
                "950": {"purity": 95.0, "grade": "Britannia Silver"},
                "999": {"purity": 99.9, "grade": "Fine Silver"},
            },
            "huid_required": True,
            "huid_format": "6 alphanumeric characters",
        },
        "validation_rules": {
            "auto_approve_confidence": 0.85,
            "auto_reject_confidence": 0.50,
            "manual_review_range": [0.50, 0.85],
        }
    }


@app.post("/validate/huid")
async def validate_huid(huid: str = Form(...)):
    """Validate HUID format."""
    import re

    cleaned = huid.upper().strip()
    cleaned = re.sub(r'[^A-Z0-9]', '', cleaned)

    is_valid = (
        len(cleaned) == 6 and
        bool(re.match(r'^[A-Z0-9]{6}$', cleaned)) and
        bool(re.search(r'[A-Z]', cleaned))
    )

    return {
        "huid": huid,
        "cleaned": cleaned,
        "valid": is_valid,
        "errors": [] if is_valid else [
            "Invalid format" if len(cleaned) != 6 else None,
            "Must contain at least one letter" if not re.search(r'[A-Z]', cleaned) else None,
        ]
    }


# =============================================================================
# ERP INTEGRATION ENDPOINTS
# =============================================================================

class ERPRegisterRequest(BaseModel):
    """Request model for ERP image registration."""
    tag_id: str
    expected_huid: str
    s3_key: str  # S3 object key where ERP uploaded the image
    s3_bucket: Optional[str] = None  # Optional, defaults to configured bucket
    callback_url: Optional[str] = None  # URL to call when processing completes
    metadata: Optional[Dict] = None  # Additional metadata from ERP


class ERPRegisterResponse(BaseModel):
    """Response model for ERP image registration."""
    status: str
    tag_id: str
    message: str
    batch_id: Optional[int] = None
    processing_status: str = "queued"


class ERPCallbackPayload(BaseModel):
    """Payload sent to ERP callback URL."""
    tag_id: str
    expected_huid: str
    actual_huid: Optional[str]
    huid_match: bool
    confidence: float
    decision: str
    purity_code: Optional[str]
    karat: Optional[str]
    rejection_reasons: List[str]
    image_url: Optional[str]
    processed_at: str


class S3EventNotification(BaseModel):
    """S3 Event Notification structure."""
    Records: Optional[List[Dict]] = None


@app.post("/api/erp/register-item", response_model=ERPRegisterResponse)
async def erp_register_item(request: ERPRegisterRequest):
    """
    ERP Integration: Register an item with expected HUID.

    Call this endpoint BEFORE uploading the image to S3.
    This creates the batch item record so the image can be processed.

    Flow:
    1. ERP calls this endpoint with tag_id and expected_huid
    2. ERP uploads image to S3 using presigned URL or direct upload
    3. ERP calls /api/erp/process-image to trigger OCR
    4. OCR system processes and calls callback_url with results
    """
    try:
        # Check if tag already exists
        existing = db.get_batch_item_by_tag(request.tag_id)
        if existing:
            return ERPRegisterResponse(
                status="exists",
                tag_id=request.tag_id,
                message="Tag ID already registered",
                batch_id=existing.batch_id,
                processing_status=existing.status.value if hasattr(existing.status, 'value') else str(existing.status)
            )

        # Create or get ERP batch (one batch per day for ERP items)
        batch_name = f"ERP_{datetime.now().strftime('%Y%m%d')}"

        # Try to find existing ERP batch for today
        all_batches = db.get_all_batches()
        erp_batch = None
        for b in all_batches:
            if b.batch_name == batch_name:
                erp_batch = b
                break

        if not erp_batch:
            # Create new batch
            batch = Batch(
                batch_name=batch_name,
                total_items=0,
                status="pending",
                metadata={"source": "erp", "auto_created": True}
            )
            batch_id = db.create_batch(batch)
        else:
            batch_id = erp_batch.id

        # Create batch item
        item_metadata = {
            "source": "erp",
            "s3_key": request.s3_key,
            "callback_url": request.callback_url,
            **(request.metadata or {})
        }

        item = BatchItem(
            batch_id=batch_id,
            tag_id=request.tag_id.strip(),
            expected_huid=request.expected_huid.strip().upper(),
            status=ProcessingStatus.PENDING,
            metadata=item_metadata
        )
        db.create_batch_item(item)

        # Update batch count
        db.update_batch_total(batch_id)

        return ERPRegisterResponse(
            status="success",
            tag_id=request.tag_id,
            message="Item registered successfully. Upload image to S3 and call /api/erp/process-image",
            batch_id=batch_id,
            processing_status="pending"
        )

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Registration failed: {str(e)}")


@app.post("/api/erp/process-image")
async def erp_process_image(
    tag_id: str = Form(...),
    s3_key: Optional[str] = Form(None),
    s3_bucket: Optional[str] = Form(None)
):
    """
    ERP Integration: Process an image that was uploaded to S3.

    Call this after:
    1. Registering the item with /api/erp/register-item
    2. Uploading the image to S3

    The system will:
    1. Download the image from S3
    2. Run OCR processing
    3. Store results in database
    4. Call the callback_url if provided during registration
    """
    # Get batch item
    batch_item = db.get_batch_item_by_tag(tag_id)
    if not batch_item:
        raise HTTPException(
            status_code=404,
            detail=f"Tag ID '{tag_id}' not found. Register it first with /api/erp/register-item"
        )

    # Get S3 key from item metadata or parameter
    item_metadata = batch_item.metadata if isinstance(batch_item.metadata, dict) else {}
    actual_s3_key = s3_key or item_metadata.get("s3_key")
    actual_bucket = s3_bucket or os.getenv('S3_BUCKET_NAME')

    if not actual_s3_key:
        raise HTTPException(
            status_code=400,
            detail="S3 key not provided and not found in registration"
        )

    try:
        start_time = time.time()

        # Update status to processing
        db.update_batch_item_status(tag_id, ProcessingStatus.PROCESSING)

        # Download image from S3
        s3 = get_s3_client()
        try:
            response = s3.get_object(Bucket=actual_bucket, Key=actual_s3_key)
            image_data = response['Body'].read()
        except ClientError as e:
            db.update_batch_item_status(tag_id, ProcessingStatus.FAILED)
            raise HTTPException(
                status_code=404,
                detail=f"Image not found in S3: {actual_s3_key}"
            )

        # Open and process image
        image = Image.open(io.BytesIO(image_data))
        if image.mode != "RGB":
            image = image.convert("RGB")

        # Store image path and generate URL
        image_url = f"https://{actual_bucket}.s3.{os.getenv('AWS_REGION', 'ap-south-1')}.amazonaws.com/{actual_s3_key}"
        db.update_batch_item_image(tag_id, actual_s3_key, image_url)

        # Process with OCR
        hallmark_info = ocr_engine_v2.extract_with_hallmark_info(image)

        # Extract actual HUID
        actual_huid = hallmark_info.huid
        expected_huid = batch_item.expected_huid

        # Compare HUIDs
        huid_match = False
        if actual_huid and expected_huid:
            huid_match = actual_huid.upper().strip() == expected_huid.upper().strip()

        # Determine decision
        decision = QCDecision.PENDING
        rejection_reasons = []

        purity_valid = hallmark_info.purity_code is not None
        huid_valid = actual_huid is not None

        if not purity_valid:
            rejection_reasons.append("missing_purity_mark")
        if not huid_valid:
            rejection_reasons.append("missing_huid")
        if actual_huid and not huid_match:
            rejection_reasons.append("huid_mismatch")

        # Decision logic
        if purity_valid and huid_valid and huid_match:
            if hallmark_info.overall_confidence >= 0.85:
                decision = QCDecision.APPROVED
            elif hallmark_info.overall_confidence >= 0.50:
                decision = QCDecision.MANUAL_REVIEW
            else:
                decision = QCDecision.REJECTED
                rejection_reasons.append("low_confidence")
        elif purity_valid and huid_valid and not huid_match:
            decision = QCDecision.REJECTED
        else:
            if hallmark_info.overall_confidence < 0.50:
                decision = QCDecision.REJECTED
            else:
                decision = QCDecision.MANUAL_REVIEW

        processing_time = int((time.time() - start_time) * 1000)

        # Create OCR result record
        ocr_result = DBOCRResult(
            batch_item_id=batch_item.id,
            tag_id=tag_id,
            expected_huid=expected_huid,
            actual_huid=actual_huid,
            huid_match=huid_match,
            purity_code=hallmark_info.purity_code,
            karat=hallmark_info.karat,
            purity_percentage=hallmark_info.purity_percentage,
            confidence=hallmark_info.overall_confidence,
            decision=decision,
            rejection_reasons=rejection_reasons,
            raw_ocr_text=" ".join([r.text for r in hallmark_info.all_results]),
            processing_time_ms=processing_time,
        )
        db.create_ocr_result(ocr_result)

        # Update batch item status
        db.update_batch_item_status(tag_id, ProcessingStatus.COMPLETED)

        # Update batch progress
        batch = db.get_batch(batch_item.batch_id)
        if batch:
            db.update_batch_progress(batch.id, batch.processed_items + 1)

        result_data = {
            "status": "success",
            "tag_id": tag_id,
            "expected_huid": expected_huid,
            "actual_huid": actual_huid,
            "huid_match": huid_match,
            "confidence": round(hallmark_info.overall_confidence, 3),
            "decision": decision.value,
            "purity_code": hallmark_info.purity_code,
            "karat": hallmark_info.karat,
            "rejection_reasons": rejection_reasons,
            "image_url": image_url,
            "processing_time_ms": processing_time
        }

        # Send callback to ERP if URL provided
        callback_url = item_metadata.get("callback_url")
        if callback_url:
            try:
                import httpx
                callback_payload = {
                    "tag_id": tag_id,
                    "expected_huid": expected_huid,
                    "actual_huid": actual_huid,
                    "huid_match": huid_match,
                    "confidence": hallmark_info.overall_confidence,
                    "decision": decision.value,
                    "purity_code": hallmark_info.purity_code,
                    "karat": hallmark_info.karat,
                    "rejection_reasons": rejection_reasons,
                    "image_url": image_url,
                    "processed_at": datetime.now().isoformat()
                }
                # Fire and forget - don't block on callback
                async with httpx.AsyncClient() as client:
                    await client.post(callback_url, json=callback_payload, timeout=10.0)
            except Exception as e:
                # Log but don't fail the request
                print(f"Callback to {callback_url} failed: {e}")

        return result_data

    except HTTPException:
        raise
    except Exception as e:
        db.update_batch_item_status(tag_id, ProcessingStatus.FAILED)
        raise HTTPException(status_code=500, detail=f"Processing failed: {str(e)}")


@app.post("/api/erp/upload-and-process")
async def erp_upload_and_process(
    file: UploadFile = File(...),
    tag_id: str = Form(...),
    expected_huid: str = Form(""),
    callback_url: Optional[str] = Form(None),
    metadata: Optional[str] = Form(None)  # JSON string
):
    """
    ERP Integration: Single endpoint to upload image and process in one call.

    This is a convenience endpoint that combines:
    1. Registration
    2. Image upload to S3
    3. OCR processing

    Use this for simpler integration where you want to send the image directly
    to this API instead of uploading to S3 first.
    """
    import json

    # Validate file type
    allowed_types = ["image/png", "image/jpeg", "image/jpg", "image/bmp", "image/webp"]
    if file.content_type not in allowed_types:
        raise HTTPException(
            status_code=400,
            detail="Invalid file type. Allowed: PNG, JPG, JPEG, BMP, WEBP"
        )

    try:
        # Parse metadata if provided
        parsed_metadata = {}
        if metadata:
            try:
                parsed_metadata = json.loads(metadata)
            except:
                pass

        # Check if tag already exists and has results
        existing = db.get_batch_item_by_tag(tag_id)
        if existing:
            existing_result = db.get_full_result_by_tag(tag_id)
            if existing_result and existing_result.get("decision"):
                return {
                    "status": "already_processed",
                    "tag_id": tag_id,
                    "message": "This tag has already been processed",
                    **existing_result
                }

        start_time = time.time()

        # Read and validate image
        contents = await file.read()
        image = Image.open(io.BytesIO(contents))
        if image.mode != "RGB":
            image = image.convert("RGB")

        # Create or get ERP batch
        batch_name = f"ERP_{datetime.now().strftime('%Y%m%d')}"
        all_batches = db.get_all_batches()
        erp_batch = None
        for b in all_batches:
            if b.batch_name == batch_name:
                erp_batch = b
                break

        if not erp_batch:
            batch = Batch(
                batch_name=batch_name,
                total_items=0,
                status="pending",
                metadata={"source": "erp", "auto_created": True}
            )
            batch_id = db.create_batch(batch)
        else:
            batch_id = erp_batch.id

        # Create batch item if doesn't exist
        if not existing:
            item_metadata = {
                "source": "erp_direct",
                "callback_url": callback_url,
                **parsed_metadata
            }
            item = BatchItem(
                batch_id=batch_id,
                tag_id=tag_id.strip(),
                expected_huid=expected_huid.strip().upper(),
                status=ProcessingStatus.PROCESSING,
                metadata=item_metadata
            )
            item_id = db.create_batch_item(item)
            db.update_batch_total(batch_id)
            batch_item = db.get_batch_item_by_tag(tag_id)
        else:
            batch_item = existing
            db.update_batch_item_status(tag_id, ProcessingStatus.PROCESSING)

        # Upload to S3
        image_path, image_url = storage.upload_image(
            contents,
            tag_id,
            file.filename,
            prefix="erp-uploads"
        )
        db.update_batch_item_image(tag_id, image_path, image_url)

        # Process with OCR
        hallmark_info = ocr_engine_v2.extract_with_hallmark_info(image)

        actual_huid = hallmark_info.huid
        expected_huid_clean = expected_huid.strip().upper()

        huid_match = False
        if actual_huid and expected_huid_clean:
            huid_match = actual_huid.upper().strip() == expected_huid_clean

        # Decision logic
        decision = QCDecision.PENDING
        rejection_reasons = []

        purity_valid = hallmark_info.purity_code is not None
        huid_valid = actual_huid is not None

        if not purity_valid:
            rejection_reasons.append("missing_purity_mark")
        if not huid_valid:
            rejection_reasons.append("missing_huid")
        if actual_huid and not huid_match:
            rejection_reasons.append("huid_mismatch")

        if purity_valid and huid_valid and huid_match:
            if hallmark_info.overall_confidence >= 0.85:
                decision = QCDecision.APPROVED
            elif hallmark_info.overall_confidence >= 0.50:
                decision = QCDecision.MANUAL_REVIEW
            else:
                decision = QCDecision.REJECTED
                rejection_reasons.append("low_confidence")
        elif purity_valid and huid_valid and not huid_match:
            decision = QCDecision.REJECTED
        else:
            if hallmark_info.overall_confidence < 0.50:
                decision = QCDecision.REJECTED
            else:
                decision = QCDecision.MANUAL_REVIEW

        processing_time = int((time.time() - start_time) * 1000)

        # Create OCR result
        ocr_result = DBOCRResult(
            batch_item_id=batch_item.id,
            tag_id=tag_id,
            expected_huid=expected_huid_clean,
            actual_huid=actual_huid,
            huid_match=huid_match,
            purity_code=hallmark_info.purity_code,
            karat=hallmark_info.karat,
            purity_percentage=hallmark_info.purity_percentage,
            confidence=hallmark_info.overall_confidence,
            decision=decision,
            rejection_reasons=rejection_reasons,
            raw_ocr_text=" ".join([r.text for r in hallmark_info.all_results]),
            processing_time_ms=processing_time,
        )
        db.create_ocr_result(ocr_result)

        db.update_batch_item_status(tag_id, ProcessingStatus.COMPLETED)

        batch = db.get_batch(batch_item.batch_id)
        if batch:
            db.update_batch_progress(batch.id, batch.processed_items + 1)

        result_data = {
            "status": "success",
            "tag_id": tag_id,
            "expected_huid": expected_huid_clean,
            "actual_huid": actual_huid,
            "huid_match": huid_match,
            "confidence": round(hallmark_info.overall_confidence, 3),
            "decision": decision.value,
            "purity_code": hallmark_info.purity_code,
            "karat": hallmark_info.karat,
            "purity_percentage": hallmark_info.purity_percentage,
            "rejection_reasons": rejection_reasons,
            "image_url": image_url,
            "processing_time_ms": processing_time,
            "message": "Image processed successfully"
        }

        # Send callback if provided
        if callback_url:
            try:
                import httpx
                callback_payload = {
                    **result_data,
                    "processed_at": datetime.now().isoformat()
                }
                async with httpx.AsyncClient() as client:
                    await client.post(callback_url, json=callback_payload, timeout=10.0)
            except Exception as e:
                print(f"Callback to {callback_url} failed: {e}")

        return result_data

    except HTTPException:
        raise
    except Exception as e:
        db.update_batch_item_status(tag_id, ProcessingStatus.FAILED)
        raise HTTPException(status_code=500, detail=f"Processing failed: {str(e)}")


@app.post("/api/erp/s3-webhook")
async def erp_s3_webhook(request: Request):
    """
    S3 Event Notification Webhook.

    Configure S3 bucket to send event notifications to this endpoint
    when images are uploaded. The system will automatically process them.

    S3 Event Configuration:
    - Event type: s3:ObjectCreated:*
    - Destination: HTTP endpoint (this URL)
    - Prefix filter: erp-images/ (optional)

    Expected S3 key format: erp-images/{tag_id}_{expected_huid}.jpg
    Example: erp-images/TAG001_AB1234.jpg
    """
    try:
        body = await request.json()

        # Handle SNS subscription confirmation
        if body.get("Type") == "SubscriptionConfirmation":
            # Auto-confirm SNS subscription
            subscribe_url = body.get("SubscribeURL")
            if subscribe_url:
                import httpx
                async with httpx.AsyncClient() as client:
                    await client.get(subscribe_url)
            return {"status": "subscription_confirmed"}

        # Handle SNS notification wrapper
        if body.get("Type") == "Notification":
            import json
            body = json.loads(body.get("Message", "{}"))

        records = body.get("Records", [])
        results = []

        for record in records:
            event_name = record.get("eventName", "")
            if not event_name.startswith("ObjectCreated"):
                continue

            s3_info = record.get("s3", {})
            bucket = s3_info.get("bucket", {}).get("name")
            key = s3_info.get("object", {}).get("key")

            if not key:
                continue

            # Parse tag_id and expected_huid from key
            # Expected format: prefix/TAG001_AB1234.jpg
            import urllib.parse
            key = urllib.parse.unquote_plus(key)

            filename = key.split("/")[-1]
            name_part = filename.rsplit(".", 1)[0]  # Remove extension

            parts = name_part.split("_", 1)
            if len(parts) != 2:
                results.append({
                    "key": key,
                    "status": "skipped",
                    "reason": "Invalid filename format. Expected: TAG_HUID.ext"
                })
                continue

            tag_id, expected_huid = parts

            # Check if already processed
            existing = db.get_batch_item_by_tag(tag_id)
            if existing:
                existing_result = db.get_full_result_by_tag(tag_id)
                if existing_result and existing_result.get("decision"):
                    results.append({
                        "tag_id": tag_id,
                        "status": "already_processed",
                        "decision": existing_result.get("decision")
                    })
                    continue

            # Register and process
            try:
                # Create ERP batch
                batch_name = f"ERP_S3_{datetime.now().strftime('%Y%m%d')}"
                all_batches = db.get_all_batches()
                erp_batch = None
                for b in all_batches:
                    if b.batch_name == batch_name:
                        erp_batch = b
                        break

                if not erp_batch:
                    batch = Batch(
                        batch_name=batch_name,
                        total_items=0,
                        status="pending",
                        metadata={"source": "s3_webhook"}
                    )
                    batch_id = db.create_batch(batch)
                else:
                    batch_id = erp_batch.id

                # Create batch item
                if not existing:
                    item = BatchItem(
                        batch_id=batch_id,
                        tag_id=tag_id,
                        expected_huid=expected_huid.upper(),
                        status=ProcessingStatus.PROCESSING,
                        metadata={"source": "s3_webhook", "s3_key": key}
                    )
                    db.create_batch_item(item)
                    db.update_batch_total(batch_id)

                batch_item = db.get_batch_item_by_tag(tag_id)

                # Download and process image
                s3 = get_s3_client()
                response = s3.get_object(Bucket=bucket, Key=key)
                image_data = response['Body'].read()

                image = Image.open(io.BytesIO(image_data))
                if image.mode != "RGB":
                    image = image.convert("RGB")

                image_url = f"https://{bucket}.s3.{os.getenv('AWS_REGION', 'ap-south-1')}.amazonaws.com/{key}"
                db.update_batch_item_image(tag_id, key, image_url)

                hallmark_info = ocr_engine_v2.extract_with_hallmark_info(image)

                actual_huid = hallmark_info.huid
                huid_match = actual_huid and actual_huid.upper() == expected_huid.upper()

                # Decision logic
                rejection_reasons = []
                if not hallmark_info.purity_code:
                    rejection_reasons.append("missing_purity_mark")
                if not actual_huid:
                    rejection_reasons.append("missing_huid")
                if actual_huid and not huid_match:
                    rejection_reasons.append("huid_mismatch")

                if hallmark_info.purity_code and actual_huid and huid_match:
                    if hallmark_info.overall_confidence >= 0.85:
                        decision = QCDecision.APPROVED
                    elif hallmark_info.overall_confidence >= 0.50:
                        decision = QCDecision.MANUAL_REVIEW
                    else:
                        decision = QCDecision.REJECTED
                        rejection_reasons.append("low_confidence")
                elif hallmark_info.purity_code and actual_huid and not huid_match:
                    decision = QCDecision.REJECTED
                else:
                    decision = QCDecision.MANUAL_REVIEW if hallmark_info.overall_confidence >= 0.50 else QCDecision.REJECTED

                ocr_result = DBOCRResult(
                    batch_item_id=batch_item.id,
                    tag_id=tag_id,
                    expected_huid=expected_huid.upper(),
                    actual_huid=actual_huid,
                    huid_match=huid_match,
                    purity_code=hallmark_info.purity_code,
                    karat=hallmark_info.karat,
                    purity_percentage=hallmark_info.purity_percentage,
                    confidence=hallmark_info.overall_confidence,
                    decision=decision,
                    rejection_reasons=rejection_reasons,
                    raw_ocr_text=" ".join([r.text for r in hallmark_info.all_results]),
                )
                db.create_ocr_result(ocr_result)
                db.update_batch_item_status(tag_id, ProcessingStatus.COMPLETED)

                results.append({
                    "tag_id": tag_id,
                    "status": "processed",
                    "decision": decision.value,
                    "huid_match": huid_match,
                    "confidence": hallmark_info.overall_confidence
                })

            except Exception as e:
                results.append({
                    "tag_id": tag_id,
                    "status": "error",
                    "error": str(e)
                })

        return {"status": "success", "processed": len(results), "results": results}

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Webhook processing failed: {str(e)}")


@app.get("/api/erp/pending-items")
async def get_erp_pending_items(
    limit: int = Query(100, le=500),
    offset: int = Query(0),
    decision: Optional[str] = Query(None, description="Filter by decision: manual_review, pending")
):
    """
    Get items pending review or action.

    Useful for dashboard to show items needing attention.
    """
    # Get all pending items from today's ERP batches
    all_batches = db.get_all_batches()
    erp_batches = [b for b in all_batches if b.batch_name.startswith("ERP_")]

    pending_items = []
    for batch in erp_batches:
        items = db.get_batch_items(batch.id)
        for item in items:
            result = db.get_full_result_by_tag(item.tag_id)

            # Filter by decision if specified
            item_decision = result.get("decision") if result else None
            if decision:
                if item_decision != decision:
                    continue
            else:
                # By default, show manual_review and pending
                if item_decision not in [None, "manual_review", "pending"]:
                    continue

            pending_items.append({
                "tag_id": item.tag_id,
                "expected_huid": item.expected_huid,
                "batch_name": batch.batch_name,
                "status": item.status.value if hasattr(item.status, 'value') else str(item.status),
                "image_url": item.image_url,
                "result": result
            })

    # Apply pagination
    total = len(pending_items)
    pending_items = pending_items[offset:offset + limit]

    return {
        "status": "success",
        "total": total,
        "offset": offset,
        "limit": limit,
        "items": pending_items
    }


@app.post("/api/erp/manual-decision")
async def set_manual_decision(
    tag_id: str = Form(...),
    decision: str = Form(..., description="approved, rejected"),
    reason: Optional[str] = Form(None),
    reviewer: Optional[str] = Form(None)
):
    """
    Manually set a decision for an item (from QC dashboard).

    Used when items are in manual_review status and need human decision.
    """
    valid_decisions = ["approved", "rejected"]
    if decision not in valid_decisions:
        raise HTTPException(
            status_code=400,
            detail=f"Invalid decision. Must be one of: {valid_decisions}"
        )

    batch_item = db.get_batch_item_by_tag(tag_id)
    if not batch_item:
        raise HTTPException(status_code=404, detail="Tag ID not found")

    # Update the OCR result with manual decision
    result = db.get_full_result_by_tag(tag_id)
    if not result:
        raise HTTPException(status_code=404, detail="No OCR result found for this tag")

    # Update decision in database
    new_decision = QCDecision.APPROVED if decision == "approved" else QCDecision.REJECTED

    rejection_reasons = result.get("rejection_reasons", [])
    if decision == "rejected" and reason:
        rejection_reasons.append(f"manual: {reason}")

    db.update_ocr_result_decision(
        tag_id=tag_id,
        decision=new_decision,
        rejection_reasons=rejection_reasons,
        reviewer=reviewer
    )

    # Get updated result
    updated_result = db.get_full_result_by_tag(tag_id)

    return {
        "status": "success",
        "tag_id": tag_id,
        "decision": decision,
        "message": f"Decision updated to {decision}",
        "result": updated_result
    }


@app.get("/api/erp/statistics")
async def get_erp_statistics(
    date: Optional[str] = Query(None, description="Date in YYYY-MM-DD format")
):
    """
    Get ERP processing statistics.

    Returns summary of items processed, decisions, and accuracy metrics.
    """
    target_date = date or datetime.now().strftime('%Y%m%d')
    batch_prefix = f"ERP_{target_date.replace('-', '')}"

    all_batches = db.get_all_batches()
    erp_batches = [b for b in all_batches if b.batch_name.startswith(batch_prefix)]

    stats = {
        "date": target_date,
        "total_items": 0,
        "processed": 0,
        "pending": 0,
        "approved": 0,
        "rejected": 0,
        "manual_review": 0,
        "huid_matches": 0,
        "huid_mismatches": 0,
        "average_confidence": 0.0,
        "batches": []
    }

    all_confidences = []

    for batch in erp_batches:
        batch_stats = db.get_batch_statistics(batch.id)
        stats["total_items"] += batch.total_items
        stats["processed"] += batch_stats.get("status_counts", {}).get("completed", 0)
        stats["pending"] += batch_stats.get("status_counts", {}).get("pending", 0)
        stats["approved"] += batch_stats.get("decision_counts", {}).get("approved", 0)
        stats["rejected"] += batch_stats.get("decision_counts", {}).get("rejected", 0)
        stats["manual_review"] += batch_stats.get("decision_counts", {}).get("manual_review", 0)
        stats["huid_matches"] += batch_stats.get("huid_matches", 0)
        stats["huid_mismatches"] += batch_stats.get("huid_mismatches", 0)

        if batch_stats.get("average_confidence"):
            all_confidences.append(batch_stats["average_confidence"])

        stats["batches"].append({
            "batch_id": batch.id,
            "batch_name": batch.batch_name,
            "total_items": batch.total_items,
            "processed_items": batch.processed_items
        })

    if all_confidences:
        stats["average_confidence"] = round(sum(all_confidences) / len(all_confidences), 3)

    return stats


if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)
